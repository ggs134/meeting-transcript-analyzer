"""
여러 회의록을 분석하여 팀원별 종합 성과표를 생성하는 스크립트
"""

import os
import re
from datetime import datetime, timedelta
from collections import defaultdict
from typing import Dict, List, Any
import json

import sys
from pathlib import Path

# 상위 디렉토리를 경로에 추가
sys.path.insert(0, str(Path(__file__).parent.parent))

from meeting_performance_analyzer import MeetingPerformanceAnalyzer


class TeamPerformanceReport:
    """팀 전체 성과 리포트 생성 클래스"""
    
    def __init__(self, analyzer: MeetingPerformanceAnalyzer):
        """
        Args:
            analyzer: 설정된 MeetingPerformanceAnalyzer 인스턴스
        """
        self.analyzer = analyzer
        self.team_data = defaultdict(lambda: {
            'total_meetings': 0,
            'total_speak_count': 0,
            'total_words': 0,
            'ideas': [],
            'action_items': [],
            'completed_tasks': [],
            'planned_tasks': [],
            'coordination': [],
            'contribution_scores': [],
            'meetings_participated': []
        })
    
    def analyze_multiple_meetings(self, filters: Dict[str, Any] = None) -> List[Dict]:
        """
        여러 회의를 분석하고 참여자별 데이터 집계
        
        Args:
            filters: MongoDB 쿼리 필터
            
        Returns:
            각 회의별 분석 결과
        """
        print("\n" + "="*70)
        print("📊 팀 성과 분석 시작")
        print("="*70)
        
        # 회의 분석 (default 템플릿 사용)
        results = self.analyzer.analyze_multiple_meetings(
            filters=filters,
            template_override="default"
        )
        
        print(f"\n✅ 총 {len(results)}개 회의 분석 완료")
        
        # 데이터 집계
        self._aggregate_data(results)
        
        return results
    
    def _aggregate_data(self, results: List[Dict]):
        """분석 결과에서 참여자별 데이터 집계"""
        
        print("\n📈 참여자별 데이터 집계 중...")
        
        for result in results:
            meeting_title = result['meeting_title']
            meeting_date = result['meeting_date']
            participants = result['participants']
            analysis = result['analysis']['analysis']
            # participant_stats는 analysis 안에 저장됨
            analysis = result.get('analysis', {})
            stats = analysis.get('participant_stats', {})
            
            # 각 참여자 데이터 추출 및 집계
            for participant in participants:
                self.team_data[participant]['total_meetings'] += 1
                self.team_data[participant]['meetings_participated'].append({
                    'title': meeting_title,
                    'date': meeting_date
                })
                
                # 통계 집계
                if participant in stats:
                    self.team_data[participant]['total_speak_count'] += stats[participant]['speak_count']
                    self.team_data[participant]['total_words'] += stats[participant]['total_words']
                
                # 분석 내용에서 정보 추출 (간단한 파싱)
                participant_section = self._extract_participant_section(analysis, participant)
                if participant_section:
                    self._parse_participant_data(participant, participant_section)
        
        print(f"✅ {len(self.team_data)}명의 팀원 데이터 집계 완료")
    
    def _extract_participant_section(self, analysis: str, participant: str) -> str:
        """분석 텍스트에서 특정 참여자 섹션 추출"""
        # [참여자 이름]으로 시작하는 섹션 찾기
        pattern = rf'\[{re.escape(participant)}\](.*?)(?=\[|---|\Z)'
        match = re.search(pattern, analysis, re.DOTALL)
        
        if match:
            return match.group(1).strip()
        return ""
    
    def _parse_participant_data(self, participant: str, section: str):
        """참여자 섹션에서 데이터 파싱"""
        # 아이디어 추출
        ideas_match = re.search(r'아이디어[\s\S]*?새로운 제안:\s*(.+?)(?=\n|$)', section)
        if ideas_match and ideas_match.group(1).strip() not in ['없음', '-', '']:
            self.team_data[participant]['ideas'].append(ideas_match.group(1).strip())
        
        # 액션 아이템 추출
        action_match = re.search(r'할일 부여:\s*(.+?)(?=\n|$)', section)
        if action_match and action_match.group(1).strip() not in ['없음', '-', '']:
            self.team_data[participant]['action_items'].append(action_match.group(1).strip())
        
        # 완료 업무 추출
        completed_match = re.search(r'수행한 일:\s*(.+?)(?=\n|$)', section)
        if completed_match and completed_match.group(1).strip() not in ['없음', '-', '']:
            self.team_data[participant]['completed_tasks'].append(completed_match.group(1).strip())
        
        # 계획 업무 추출
        planned_match = re.search(r'수행할 일:\s*(.+?)(?=\n|$)', section)
        if planned_match and planned_match.group(1).strip() not in ['없음', '-', '']:
            self.team_data[participant]['planned_tasks'].append(planned_match.group(1).strip())
        
        # 기여도 점수 추출
        score_match = re.search(r'양적 기여도:\s*(\d+)', section)
        if score_match:
            self.team_data[participant]['contribution_scores'].append(int(score_match.group(1)))
    
    def generate_summary_report(self) -> str:
        """팀 전체 요약 리포트 생성"""
        
        report = []
        report.append("="*70)
        report.append("📊 팀 전체 성과 요약 리포트")
        report.append("="*70)
        report.append(f"\n생성 일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append(f"분석 팀원 수: {len(self.team_data)}명")
        
        # 참여자별 요약
        report.append("\n\n" + "="*70)
        report.append("👥 참여자별 요약")
        report.append("="*70)
        
        # 평균 기여도 순으로 정렬
        sorted_participants = sorted(
            self.team_data.items(),
            key=lambda x: sum(x[1]['contribution_scores']) / len(x[1]['contribution_scores']) if x[1]['contribution_scores'] else 0,
            reverse=True
        )
        
        for participant, data in sorted_participants:
            report.append(f"\n{'─'*70}")
            report.append(f"👤 {participant}")
            report.append(f"{'─'*70}")
            
            # 기본 통계
            report.append(f"\n📊 참여 통계:")
            report.append(f"  • 참여 회의: {data['total_meetings']}회")
            report.append(f"  • 총 발언 횟수: {data['total_speak_count']}회")
            report.append(f"  • 총 발언 단어 수: {data['total_words']}개")
            report.append(f"  • 회의당 평균 발언: {data['total_speak_count'] / data['total_meetings']:.1f}회" if data['total_meetings'] > 0 else "  • 회의당 평균 발언: 0회")
            
            # 평균 기여도
            if data['contribution_scores']:
                avg_contribution = sum(data['contribution_scores']) / len(data['contribution_scores'])
                report.append(f"  • 평균 기여도: {avg_contribution:.1f}/100")
            
            # 아이디어
            if data['ideas']:
                report.append(f"\n💡 제안한 아이디어: {len(data['ideas'])}개")
                for i, idea in enumerate(data['ideas'][:5], 1):  # 최대 5개만
                    report.append(f"  {i}. {idea[:100]}...")
                if len(data['ideas']) > 5:
                    report.append(f"  ... 외 {len(data['ideas']) - 5}개")
            
            # 완료 업무
            if data['completed_tasks']:
                report.append(f"\n✅ 완료 보고: {len(data['completed_tasks'])}건")
                for i, task in enumerate(data['completed_tasks'][:3], 1):
                    report.append(f"  {i}. {task[:100]}...")
            
            # 계획 업무
            if data['planned_tasks']:
                report.append(f"\n📋 수행 예정: {len(data['planned_tasks'])}건")
                for i, task in enumerate(data['planned_tasks'][:3], 1):
                    report.append(f"  {i}. {task[:100]}...")
            
            # 액션 아이템
            if data['action_items']:
                report.append(f"\n🎯 조율한 업무: {len(data['action_items'])}건")
        
        # 전체 통계
        report.append("\n\n" + "="*70)
        report.append("📈 전체 통계")
        report.append("="*70)
        
        total_meetings = len(set(
            meeting['title'] 
            for data in self.team_data.values() 
            for meeting in data['meetings_participated']
        ))
        
        report.append(f"\n총 분석 회의: {total_meetings}개")
        report.append(f"총 참여자: {len(self.team_data)}명")
        report.append(f"총 발언 횟수: {sum(d['total_speak_count'] for d in self.team_data.values())}회")
        report.append(f"총 단어 수: {sum(d['total_words'] for d in self.team_data.values())}개")
        
        return "\n".join(report)
    
    def generate_detailed_table(self) -> str:
        """상세 성과표 생성 (표 형태)"""
        
        table = []
        table.append("="*150)
        table.append("📊 팀원별 상세 성과표")
        table.append("="*150)
        
        # 헤더
        header = f"{'이름':<10} {'회의참여':<8} {'발언횟수':<8} {'단어수':<8} {'평균기여도':<10} {'아이디어':<8} {'완료':<6} {'계획':<6} {'조율':<6}"
        table.append(header)
        table.append("-"*150)
        
        # 평균 기여도 순으로 정렬
        sorted_participants = sorted(
            self.team_data.items(),
            key=lambda x: sum(x[1]['contribution_scores']) / len(x[1]['contribution_scores']) if x[1]['contribution_scores'] else 0,
            reverse=True
        )
        
        # 데이터 행
        for participant, data in sorted_participants:
            avg_contribution = sum(data['contribution_scores']) / len(data['contribution_scores']) if data['contribution_scores'] else 0
            
            row = (
                f"{participant:<10} "
                f"{data['total_meetings']:<8} "
                f"{data['total_speak_count']:<8} "
                f"{data['total_words']:<8} "
                f"{avg_contribution:<10.1f} "
                f"{len(data['ideas']):<8} "
                f"{len(data['completed_tasks']):<6} "
                f"{len(data['planned_tasks']):<6} "
                f"{len(data['action_items']):<6}"
            )
            table.append(row)
        
        table.append("="*150)
        
        return "\n".join(table)
    
    def export_to_json(self, filename: str = "team_performance.json"):
        """JSON 파일로 내보내기"""
        
        output_path = f"/mnt/user-data/outputs/{filename}"
        
        export_data = {
            'generated_at': datetime.now().isoformat(),
            'total_participants': len(self.team_data),
            'participants': {}
        }
        
        for participant, data in self.team_data.items():
            export_data['participants'][participant] = {
                'total_meetings': data['total_meetings'],
                'total_speak_count': data['total_speak_count'],
                'total_words': data['total_words'],
                'avg_contribution_score': sum(data['contribution_scores']) / len(data['contribution_scores']) if data['contribution_scores'] else 0,
                'ideas_count': len(data['ideas']),
                'completed_tasks_count': len(data['completed_tasks']),
                'planned_tasks_count': len(data['planned_tasks']),
                'action_items_count': len(data['action_items']),
                'meetings_participated': data['meetings_participated'],
                'ideas': data['ideas'],
                'completed_tasks': data['completed_tasks'],
                'planned_tasks': data['planned_tasks'],
                'action_items': data['action_items'],
                'contribution_scores': data['contribution_scores']
            }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, ensure_ascii=False, indent=2)
        
        print(f"\n💾 JSON 파일 저장: {output_path}")
        return output_path
    
    def export_to_csv(self, filename: str = "team_performance.csv"):
        """CSV 파일로 내보내기 (Excel에서 열기 가능)"""
        
        output_path = f"/mnt/user-data/outputs/{filename}"
        
        lines = []
        
        # 헤더
        headers = [
            "이름", "참여회의수", "총발언횟수", "총단어수", 
            "평균기여도", "아이디어수", "완료업무수", "계획업무수", "조율업무수"
        ]
        lines.append(",".join(headers))
        
        # 데이터
        sorted_participants = sorted(
            self.team_data.items(),
            key=lambda x: sum(x[1]['contribution_scores']) / len(x[1]['contribution_scores']) if x[1]['contribution_scores'] else 0,
            reverse=True
        )
        
        for participant, data in sorted_participants:
            avg_contribution = sum(data['contribution_scores']) / len(data['contribution_scores']) if data['contribution_scores'] else 0
            
            row = [
                participant,
                str(data['total_meetings']),
                str(data['total_speak_count']),
                str(data['total_words']),
                f"{avg_contribution:.1f}",
                str(len(data['ideas'])),
                str(len(data['completed_tasks'])),
                str(len(data['planned_tasks'])),
                str(len(data['action_items']))
            ]
            lines.append(",".join(row))
        
        with open(output_path, 'w', encoding='utf-8-sig') as f:  # BOM 추가로 Excel 호환성
            f.write("\n".join(lines))
        
        print(f"💾 CSV 파일 저장: {output_path}")
        return output_path
    
    def export_to_excel(self, filename: str = "team_performance.xlsx"):
        """Excel 파일로 내보내기 (차트 포함)"""
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import BarChart, Reference
        except ImportError:
            print("⚠️  openpyxl이 설치되지 않았습니다. CSV로만 저장합니다.")
            return None
        
        output_path = f"/mnt/user-data/outputs/{filename}"
        
        wb = Workbook()
        
        # 시트 1: 요약 표
        ws1 = wb.active
        ws1.title = "팀 성과 요약"
        
        # 헤더 스타일
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 헤더
        headers = [
            "이름", "참여회의", "발언횟수", "단어수", 
            "평균기여도", "아이디어", "완료", "계획", "조율"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # 데이터
        sorted_participants = sorted(
            self.team_data.items(),
            key=lambda x: sum(x[1]['contribution_scores']) / len(x[1]['contribution_scores']) if x[1]['contribution_scores'] else 0,
            reverse=True
        )
        
        for row_idx, (participant, data) in enumerate(sorted_participants, 2):
            avg_contribution = sum(data['contribution_scores']) / len(data['contribution_scores']) if data['contribution_scores'] else 0
            
            row_data = [
                participant,
                data['total_meetings'],
                data['total_speak_count'],
                data['total_words'],
                round(avg_contribution, 1),
                len(data['ideas']),
                len(data['completed_tasks']),
                len(data['planned_tasks']),
                len(data['action_items'])
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_idx > 1:  # 숫자 컬럼은 오른쪽 정렬
                    cell.alignment = Alignment(horizontal='right')
        
        # 열 너비 조정
        ws1.column_dimensions['A'].width = 12
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            ws1.column_dimensions[col].width = 12
        
        # 시트 2: 상세 정보
        ws2 = wb.create_sheet("상세 정보")
        
        ws2['A1'] = "참여자별 상세 성과"
        ws2['A1'].font = Font(bold=True, size=14)
        
        current_row = 3
        for participant, data in sorted_participants:
            # 참여자 이름
            ws2[f'A{current_row}'] = participant
            ws2[f'A{current_row}'].font = Font(bold=True, size=12)
            ws2[f'A{current_row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            current_row += 1
            
            # 통계
            ws2[f'A{current_row}'] = "참여 회의"
            ws2[f'B{current_row}'] = f"{data['total_meetings']}회"
            current_row += 1
            
            ws2[f'A{current_row}'] = "아이디어 수"
            ws2[f'B{current_row}'] = f"{len(data['ideas'])}개"
            current_row += 1
            
            # 아이디어 목록
            if data['ideas']:
                ws2[f'A{current_row}'] = "제안 아이디어:"
                ws2[f'A{current_row}'].font = Font(italic=True)
                current_row += 1
                
                for i, idea in enumerate(data['ideas'][:5], 1):
                    ws2[f'B{current_row}'] = f"{i}. {idea[:80]}..."
                    current_row += 1
            
            current_row += 2  # 다음 참여자와 간격
        
        # 차트 추가 (시트 3)
        ws3 = wb.create_sheet("차트")
        
        # 기여도 차트
        chart = BarChart()
        chart.title = "팀원별 평균 기여도"
        chart.x_axis.title = "팀원"
        chart.y_axis.title = "기여도"
        
        # 데이터 범위 설정
        data_range = Reference(ws1, min_col=5, min_row=1, max_row=len(sorted_participants)+1)
        categories = Reference(ws1, min_col=1, min_row=2, max_row=len(sorted_participants)+1)
        
        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories)
        
        ws3.add_chart(chart, "A1")
        
        # 저장
        wb.save(output_path)
        
        print(f"📊 Excel 파일 저장: {output_path}")
        return output_path
    
    def save_full_report(self, filename: str = "team_performance_report.txt"):
        """전체 리포트를 텍스트 파일로 저장"""
        
        output_path = f"/mnt/user-data/outputs/{filename}"
        
        with open(output_path, 'w', encoding='utf-8') as f:
            # 요약 리포트
            f.write(self.generate_summary_report())
            f.write("\n\n")
            
            # 상세 표
            f.write(self.generate_detailed_table())
        
        print(f"💾 전체 리포트 저장: {output_path}")
        return output_path


def main():
    """메인 실행 함수"""
    
    # 환경 변수에서 설정 읽기
    MONGODB_URI = os.getenv('MONGODB_URI', 'mongodb://localhost:27017/')
    DATABASE_NAME = os.getenv('DATABASE_NAME', 'company_db')
    COLLECTION_NAME = os.getenv('COLLECTION_NAME', 'meeting_transcripts')
    GEMINI_API_KEY = os.getenv('GEMINI_API_KEY', 'your-gemini-api-key-here')
    
    print("\n" + "="*70)
    print("🚀 팀 성과 리포트 생성기")
    print("="*70)
    print(f"\nDatabase: {DATABASE_NAME}")
    print(f"Collection: {COLLECTION_NAME}")
    
    # 분석기 초기화 (기본 템플릿 사용)
    analyzer = MeetingPerformanceAnalyzer(
        mongodb_uri=MONGODB_URI,
        database_name=DATABASE_NAME,
        collection_name=COLLECTION_NAME,
        gemini_api_key=GEMINI_API_KEY,
        prompt_template="default"  # 기본 템플릿으로 일관성 있게
    )
    
    # 리포트 생성기 초기화
    report_generator = TeamPerformanceReport(analyzer)
    
    # 필터 설정 (예: 최근 3개월)
    # filters = {
    #     'date': {'$gte': datetime.now() - timedelta(days=90)}
    # }
    
    # 모든 회의 분석
    filters = {}
    
    try:
        # 회의 분석 및 데이터 집계
        results = report_generator.analyze_multiple_meetings(filters)
        
        # 콘솔에 출력
        print("\n" + report_generator.generate_summary_report())
        print("\n" + report_generator.generate_detailed_table())
        
        # 파일로 저장
        print("\n\n" + "="*70)
        print("💾 리포트 파일 생성 중...")
        print("="*70)
        
        report_generator.save_full_report()
        report_generator.export_to_json()
        report_generator.export_to_csv()
        report_generator.export_to_excel()
        
        print("\n✅ 모든 리포트 생성 완료!")
        print("\n📁 생성된 파일:")
        print("  • team_performance_report.txt - 전체 리포트")
        print("  • team_performance.json - JSON 데이터")
        print("  • team_performance.csv - CSV 파일")
        print("  • team_performance.xlsx - Excel 파일 (차트 포함)")
        
    except Exception as e:
        print(f"\n❌ 오류 발생: {str(e)}")
        import traceback
        traceback.print_exc()
    
    finally:
        # 연결 종료
        analyzer.close()


if __name__ == "__main__":
    main()
